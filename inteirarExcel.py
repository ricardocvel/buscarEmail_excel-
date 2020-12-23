#encoding utf-8

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

import os
import re

def salvar_email():

    path = 'E:/4 - ARQUIVO\PROJETOS\motor\email.xlsx'
    arquivo_excel = load_workbook(path)
    separados = arquivo_excel.active  # Le a planilha ativa 
    '''
    # obter sheets
    sheets = arquivo_excel.sheetnames
    planilha1 = arquivo_excel[sheets[n]]
    '''
    #ler linha a linha
    max_linha = separados.max_row
    max_coluna = separados.max_column

    contador = 1

    for i in range(1, max_linha):

        a1 = str(separados.cell(row=i, column=1, ).value)
        #print(type(str(a1.value)))
        #print(str(a1.value))

        if re.search('\\SMTP\\b', a1, re.IGNORECASE):
            email = a1.split(',')

            for g in range(len(email)):
                if re.search('ricardo.campos', email[g], re.IGNORECASE):
                    email[g] = "0"
                if re.search('postmaster', email[g], re.IGNORECASE):
                    email[g] = "0"
                if re.search('@', email[g], re.IGNORECASE):
                    email[g] = email[g][:-1]
                    email[g] = email[g][1:]
                     # setando nova planilha para gravação (OBS não desativa a planilha ativa)
                    result = arquivo_excel['result'] 
                    result.cell(row=contador, column=2).value = email[g]
                    contador = contador + 1
                    print(email[g])

    arquivo_excel .save(path)

    os.system("PAUSE")


def retira():
    path = 'E:/4 - ARQUIVO\PROJETOS\motor\email.xlsx'
    arquivo_excel = load_workbook(path)

    # obter sheets B 01 -Intelbras
    sheets = arquivo_excel.sheetnames
    print(sheets)

    sheet1 = arquivo_excel[sheets[0]]
    sheet2 = arquivo_excel[sheets[1]]
    sheet3 = arquivo_excel[sheets[2]]
    # ler linha a linha
    max_linha = sheet3.max_row
    max_coluna = sheet3.max_column

    contador = 1

    for i in range(1, max_linha):
        a1 = str(sheet3.cell(row=i, column=2, ).value)
        #a1 = a1.split(' ')
        print(a1[5:])

        sheet2.cell(row=contador, column=4).value = a1[5:]
        contador = contador + 1
    arquivo_excel.save(path)

    os.system("PAUSE")


#salvar_email()
retira()

